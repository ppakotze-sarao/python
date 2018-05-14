#Opens cls workbook, selects particular sheet, searches for rows which start with a datetime and writes only those rows to a csv file
from openpyxl import load_workbook
from openpyxl import Workbook

import csv
import sys
import datetime

wb2_name = sys.argv[1]  #"SACAAlogbook_22032018.xlsx")

wb2 = load_workbook(wb2_name)

wb_new = Workbook()
ws=wb_new.active

print(wb2.get_sheet_names() )

worksheet1 = wb2['Pages']

# print(worksheet1['D18'].value )

your_csv_file = open('logbook.csv','wb')
wr = csv.writer(your_csv_file, quoting=csv.QUOTE_ALL)
for row in worksheet1.iter_rows():
   # print( row[0], type(row[0].value) )
   if  type(row[0].value ) == datetime.datetime: #only look for rows starting with date
      lrow = []
      # :wprint(row)
      for cell in row: 
        if type(cell.value) == unicode:
		lrow.append(cell.value.encode('utf-8')) #force unicode to utf-8
	else:
		lrow.append((cell.value))
      # print(lrow)
      wr.writerow(lrow) #need to check if csv writer is utf-8?
print('Done')
your_csv_file.close()   

